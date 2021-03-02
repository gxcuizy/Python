#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
openpyxl模块处理表格数据，在当前表格行中增加数据并保存到原表格
author: gxcuizy
time: 2021-03-02
"""

import openpyxl
from openpyxl import load_workbook
import pymysql


def get_mysql_info(cursor, order_code):
    """获取数据库数据"""
    # 组装数据sql
    sql = "SELECT order_total, order_status FROM order WHERE order_code = '%s'" % (order_code,)
    # 获取一条数据
    cursor.execute(sql)
    result = cursor.fetchone()
    return result


# 程序主入口
if __name__ == "__main__":
    # 创建一个连接
    db_host = '127.0.0.1'
    db_name = 'test'
    db_user = 'root'
    db_pw = 'root'
    db_port = 3306
    db = pymysql.connect(host=db_host, user=db_user, password=db_pw, database=db_name, port=db_port)
    # 用cursor()创建一个游标对象
    cursor_obj = db.cursor(cursor=pymysql.cursors.DictCursor)
    # 读取excel数据
    print('开始读取表格数据……')
    file_name = '1.xlsx'
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.active
    # 写表excel数据
    print('读取数据完毕，开始处理数据……')
    write_wb = openpyxl.Workbook()
    write_sheet = write_wb.active
    # 循环处理Excel数据（假设保留原先的1-3列数据，在第4列和第5列增加订单金额和状态）
    for row in sheet.rows:
        # 保留原来行数据
        row_1 = row[0].value
        row_2 = row[1].value
        row_3 = row[2].value
        # 获取附加的行数据
        print('正在处理order_code=%s' % row_1)
        info = get_mysql_info(cursor_obj, row_1)
        row_4 = ''
        row_5 = ''
        if info:
            row_4 = info['order_total']
            val_5 = info['order_status']
        excel_row = [
            row_1,
            row_2,
            row_3,
            row_4,
            row_5,
        ]
        # 写入Excel中
        write_sheet.append(excel_row)
    # 生成excel文件
    write_wb.save(file_name)
    print('写入excel完成！')

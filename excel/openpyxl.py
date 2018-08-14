#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook

# 程序主入口
if __name__ == "__main__":
    # 读取excel数据
    wb = load_workbook('test.xlsx', data_only=True)
    sheet = wb.active
    # 写表excel数据
    write_wb = openpyxl.Workbook()
    write_sheet = write_wb.active
    print('开始写入excel')
    # 表头信息
    cxcel_title = [
        'title_1',
        'title_2',
        'title_3',
        'title_4',
        'title_5',
        'title_6'
    ]
    # 表头写入Excel中
    write_sheet.append(cxcel_title)
    # 循环写入旧Excel数据
    for row in sheet.rows:
        # 行信息
        title_1 = row[0].value
        title_2 = row[1].value
        title_3 = row[2].value
        title_4 = row[3].value
        title_5 = row[4].value
        title_6 = row[5].value
        excel_row = [
            title_1,
            title_2,
            title_3,
            title_4,
            title_5,
            title_6
        ]
        # 写入Excel中
        write_sheet.append(excel_row)
    # 生成excel文件
    write_wb.save('new_test.xlsx')
    print('写入excel完成！')

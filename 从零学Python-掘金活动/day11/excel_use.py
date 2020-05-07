#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel的读写使用
author: gxcuizy
date: 2018-10-25
"""

import openpyxl


def write_excel(file):
    """创建Excel表操作，逐个单元格写值"""
    print('正在创建表格……')
    wb = openpyxl.Workbook()
    # 获取当前默认工作表
    sheet = wb.active
    # 修改工作表名称
    sheet.title = 'test'
    # 创建一个WorkSheet
    new_sheet = wb.create_sheet('new sheet')
    print('给单元格赋值……')
    # 通过坐标给单元格赋值
    sheet['A1'] = 666
    sheet['B1'] = 888
    sheet['A2'] = 999
    new_sheet['A1'] = 888888
    new_sheet['A2'] = 666666
    new_sheet['A3'] = 111111
    # 保存并生成
    wb.save(file)
    print('保存数据，写表完毕！')


def create_excel(file):
    """创建Excel表操作，批量给单元格赋值"""
    print('正在创建表格……')
    wb = openpyxl.Workbook()
    # 获取当前默认工作表
    sheet = wb.active
    # 写入标题
    print('正在写表格标题……')
    title = ['序号', '标题', '内容']
    sheet.append(title)
    # 写入主内容
    print('正在写表格主要内容……')
    excel_list_1 = ['1', '学习Python', 'hello world']
    excel_list_2 = ['2', '人生苦短', '我学Python']
    sheet.append(excel_list_1)
    sheet.append(excel_list_2)
    # 保存并生成
    wb.save(file)
    print('保存数据，写表完毕！')


def read_excel(file):
    """读取Excel操作"""
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    # rows: 返回所有有效数据行,有数据时类型为generator,无数据时为tuple
    print('-------------------------每行的数据------------------------')
    for rows in sheet.rows:
        # 循环每行的数据
        row_list = []
        for row in rows:
            row_list.append(row.value)
        print(row_list)

    # rows: 返回所有有效数据行,有数据时类型为generator,无数据时为tuple
    print('-------------------------每列的数据------------------------')
    for columns in sheet.columns:
        # 循环每列的数据
        column_list = []
        for column in columns:
            column_list.append(column.value)
        print(column_list)


def cell_excel(file):
    """已存在文件，读取以及设置单个单元格的值"""
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    # 读取单元格的值
    print('单元格B2的值为：' + sheet['B2'].value)
    # 修改单元格的值，重新生成新的表格
    sheet['B2'].value = '打死不学Python'
    wb.save('new.xlsx')


# 程序主入口
if __name__ == '__main__':
    # 写xlsx表格数据
    write_excel('write.xlsx')
    create_excel('create.xlsx')
    # 读取Excel数据
    read_excel('create.xlsx')
    # 读取单个单元格的值，并修改单元格
    cell_excel('create.xlsx')
